#!/usr/bin/env python3
# Gets the sheet names of an excel file

import xlrd
import openpyxl
import sys


def getSheetName_xlsx(file_name):
    names = []
    Workbook = openpyxl.load_workbook(file_name)
    names = Workbook.sheetnames

    for name in names:
        sheet = Workbook[name]
        if sheet.max_row != 0:
            if 'Sheet' not in name:
                print(name)
            else:
                print("")


def getSheetName_xls(file_name):
    names = []
    Workbook = xlrd.open_workbook(file_name, encoding_override="cp1252")
    names = Workbook.sheet_names()

    for name in names:
        sheet = Workbook.sheet_by_name(name)
        if sheet.nrows != 0:
            if 'Sheet' not in name:
                print(name)
            else:
                print("")


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print(" Usage: %s original.csv > entries.csv" % sys.argv[0])
        sys.exit(1)
    if 'xlsx' in sys.argv[1]:
        sys.exit(getSheetName_xlsx(sys.argv[1]))
    else:
        sys.exit(getSheetName_xls(sys.argv[1]))
