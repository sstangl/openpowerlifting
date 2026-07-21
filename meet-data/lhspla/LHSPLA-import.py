#!/usr/bin/env python3
"""
LHSPA Meet Results Transformer
=====================================

Reads an LHSPA meet workbook (.xlsx) and produces:

    original.csv
    entries.csv

The script implements the canonical LHSPA protocol.

Dependencies:
    openpyxl

Author:
    ChatGPT
"""

import csv
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


# ==========================================================
# CONSTANTS
# ==========================================================

OUTPUT_COLUMNS = [

    "WeightClassLbs",
    "Name",
    "Team",
    "Division",
    "BodyweightLbs",
    "Best3SquatLbs",
    "Best3BenchLbs",
    "Best3DeadliftLbs",
    "TotalLbs",
    "Place",
    "Sex",
    "Event",
    "Equipment",
    "BirthDate"

]


HEADER_MAP = {

    "WeightClassLbs": [
        "WT",
        "WT CLASS",
        "WT. CLASS",
        "WEIGHT CLASS",
        "WEIGHT CLASS LBS",
        "CLASS"
    ],

    "Name": [
        "NAME",
        "LIFTER",
        "LIFTER NAME"
    ],

    "Team": [
        "TEAM",
        "SCHOOL"
    ],

    "BodyweightLbs": [
        "BODY WEIGHT",
        "BODY WEIGHT LBS",
        "BODYWEIGHT",
        "BODYWT",
        "BWT"
    ],

    "Best3SquatLbs": [
        "SQUAT",
        "BEST SQUAT"
    ],

    "Best3BenchLbs": [
        "BENCH",
        "BENCH PRESS",
        "BEST BENCH"
    ],

    "Best3DeadliftLbs": [
        "DEAD LIFT",
        "DEADLIFT",
        "BEST DEADLIFT"
    ],

    "TotalLbs": [
        "TOTAL"
    ],

    "Place": [
        "PLACE",
        "PLACING"
    ]

}


# ==========================================================
# GENERAL UTILITIES
# ==========================================================

def log(message):
    """Display progress."""

    print(f"[LHSPA] {message}")


def error(message):
    """Display an error and terminate."""

    print()
    print("=" * 60)
    print("ERROR")
    print("=" * 60)
    print(message)
    sys.exit(1)


def clean_text(value):
    """
    Convert None to blank and trim whitespace.
    """

    if value is None:
        return ""

    return str(value).strip()


def normalize_header(value):
    """
    Make headers easy to compare.

    Example:

        Wt. Class

    becomes

        WT CLASS
    """

    value = clean_text(value).upper()

    value = value.replace(".", " ")

    value = re.sub(r"\s+", " ", value)

    return value.strip()


def is_blank_row(row):
    """
    True if every cell is empty.
    """

    for value in row:

        if clean_text(value) != "":
            return False

    return True


# ==========================================================
# NUMBER HANDLING
# ==========================================================

def remove_zero(value):
    """
    Replace numeric zero with blank.
    """

    if value is None:
        return ""

    try:

        number = float(value)

        if number == 0:
            return ""

        return value

    except Exception:

        return value


def format_number(value):
    """
    Format values for CSV.

    181.0  -> 181
    181.5  -> 181.5
    0      -> blank
    """

    value = remove_zero(value)

    if value == "":
        return ""

    try:

        number = float(value)

        if number.is_integer():
            return str(int(number))

        return str(number)

    except Exception:

        return str(value).strip()


# ==========================================================
# NAME TRANSFORMATIONS
# ==========================================================

def transform_name(name):
    """
    Transform names into LHSPA format.

    Examples:

        SMITH, JOHN 4
            -> John Smith

        SMITH, JOHN 12
            -> John Smith

        "SMITH, JOHN 7"
            -> John Smith

        SMITH, JOHN
            -> John Smith
    """

    name = clean_text(name)

    #
    # Remove quotation marks
    #
    name = name.replace('"', "")

    #
    # Remove trailing one or two digit competitor numbers
    #
    name = re.sub(r"(?<=[A-Za-z])\d{1,2}(?=\s|$)", "", name)

    #
    # Convert LAST, FIRST -> FIRST LAST
    #
    if "," in name:

        last, first = name.split(",", 1)

        name = f"{first.strip()} {last.strip()}"

    #
    # Collapse multiple spaces
    #
    name = re.sub(r"\s+", " ", name).strip()

    #
    # Convert ALL CAPS to Proper Case
    #
    name = name.title()

    return name


# ==========================================================
# FILENAME DERIVED VALUES
# ==========================================================

def determine_division(filename):

    filename = filename.upper()

    if "GIRL" in filename:
        return "Girls"

    if "BOY" in filename:
        return "Boys"

    return ""


def determine_sex(filename):

    filename = filename.upper()

    if "GIRL" in filename:
        return "F"

    if "BOY" in filename:
        return "M"

    return ""


# ==========================================================
# FIELD TRANSFORMATIONS
# ==========================================================

def transform_weight_class(weight_class, division):

    weight_class = clean_text(weight_class)

    if weight_class.upper() == "SHW":

        if division == "Girls":
            return "220+"

        return "275+"

    return weight_class


def transform_place(place):

    place = clean_text(place).upper()

    if place == "":
        return "DQ"

    if place == "BO":
        return "DQ"

    return place


def transform_cell(value):
    """
    Apply generic transformations.
    """

    value = remove_zero(value)

    if value == "":
        return ""

    return value
# ==========================================================
# WORKBOOK READER
# ==========================================================

def read_workbook(filename):
    """
    Open the workbook and return the first worksheet.
    """

    log(f"Opening workbook: {filename}")

    workbook = load_workbook(
        filename,
        data_only=True
    )

    worksheet = workbook[workbook.sheetnames[0]]

    return workbook, worksheet


def worksheet_to_rows(worksheet):
    """
    Read every worksheet row.

    Numeric cells with strikethrough formatting are converted
    to negative values.
    """

    rows = []

    for excel_row in worksheet.iter_rows():

        row = []

        for cell in excel_row:

            value = cell.value

            if value is not None:

                try:

                    if (
                        isinstance(value, (int, float))
                        and cell.font is not None
                        and cell.font.strike
                    ):
                        value = -abs(value)

                except Exception:
                    pass

            row.append(value)

        rows.append(row)

    return rows


# ==========================================================
# ORIGINAL.CSV
# ==========================================================

def write_original_csv(rows, output_folder):
    """
    Write original.csv preserving every row exactly
    as read from the workbook.
    """

    output_file = output_folder / "original.csv"

    with open(
        output_file,
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as csvfile:

        writer = csv.writer(csvfile)

        for row in rows:
            writer.writerow(row)

    log(f"Created {output_file}")


# ==========================================================
# TABLE DETECTION
# ==========================================================

def find_header_row(rows):
    """
    Find the header row by looking for the weight
    class column.
    """

    for row_index, row in enumerate(rows):

        headers = [
            normalize_header(cell)
            for cell in row
        ]

        if (
            "WT CLASS" in headers
            or "WT. CLASS" in headers
            or "WEIGHT CLASS" in headers
            or "WT" in headers
        ):

            log(f"Header row found at row {row_index + 1}")

            return row_index

    error(
        "Unable to locate the results table header."
    )


def find_end_of_first_table(rows, header_row):
    """
    The first completely blank row marks the end
    of the individual results table.
    """

    for row_index in range(header_row + 1, len(rows)):

        if is_blank_row(rows[row_index]):

            log(
                f"First table ends at row {row_index}"
            )

            return row_index

    return len(rows)


def extract_first_table(rows):
    """
    Return only the first results table.
    """

    header_row = find_header_row(rows)

    end_row = find_end_of_first_table(
        rows,
        header_row
    )

    return (
        rows[header_row],
        rows[header_row + 1:end_row]
    )


# ==========================================================
# FLEXIBLE COLUMN DETECTION
# ==========================================================

def locate_columns(header_row):
    """
    Match worksheet columns to LHSPA output fields.

    Matching is case-insensitive and tolerant of
    punctuation and extra text.

    Example:
        "Body Weight (lbs)"
        "Body Weight"
        "BWT"

    all map to BodyweightLbs.
    """

    columns = {}

    normalized_headers = [
        normalize_header(cell)
        for cell in header_row
    ]

    for target, aliases in HEADER_MAP.items():

        found = False

        for column_number, header in enumerate(normalized_headers):

            for alias in aliases:

                alias = normalize_header(alias)

                #
                # Exact match
                #

                if header == alias:

                    columns[target] = column_number
                    found = True
                    break

                #
                # Header contains alias
                #

                if alias in header:

                    columns[target] = column_number
                    found = True
                    break

                #
                # Alias contains header
                #

                if header in alias:

                    columns[target] = column_number
                    found = True
                    break

            if found:
                break

        if not found:

            print()
            print("Worksheet headers found:\n")

            for header in normalized_headers:
                print(f"  {header}")

            error(
                f'Unable to locate required column "{target}".'
            )

    log("Column mapping complete.")

    return columns
# ==========================================================
# ENTRY TRANSFORMATION
# ==========================================================

def build_entry(row, columns, division, sex):
    """
    Convert one worksheet row into one LHSPA entry.
    """

    def get(field):
        index = columns[field]

        if index >= len(row):
            return ""

        return row[index]

    weight_class = transform_weight_class(
        get("WeightClassLbs"),
        division
    )

    bodyweight = format_number(
        get("BodyweightLbs")
    )

    squat = format_number(
        get("Best3SquatLbs")
    )

    bench = format_number(
        get("Best3BenchLbs")
    )

    deadlift = format_number(
        get("Best3DeadliftLbs")
    )

    total = format_number(
        get("TotalLbs")
    )

    place = transform_place(
        get("Place")
    )

    return {

        "WeightClassLbs": weight_class,

        "Name": transform_name(
            get("Name")
        ),

        "Team": clean_text(
            get("Team")
        ),

        "Division": division,

        "BodyweightLbs": bodyweight,

        "Best3SquatLbs": squat,

        "Best3BenchLbs": bench,

        "Best3DeadliftLbs": deadlift,

        "TotalLbs": total,

        "Place": place,

        "Sex": sex,

        "Event": "SBD",

        "Equipment": "Single-ply",

        "BirthDate": ""

    }


# ==========================================================
# BUILD ENTRIES
# ==========================================================

def build_entries(table_header,
                  table_rows,
                  filename):
    """
    Convert the worksheet into LHSPA entries.
    """

    division = determine_division(filename)

    sex = determine_sex(filename)

    columns = locate_columns(table_header)

    entries = []

    skipped = 0

    for row in table_rows:

        #
        # Ignore completely blank rows
        #

        if is_blank_row(row):
            continue

        #
        # Ignore rows without a lifter name
        #

        try:

            name_index = columns["Name"]

            if clean_text(row[name_index]) == "":
                skipped += 1
                continue

        except Exception:

            skipped += 1
            continue

        #
        # Build the entry
        #

        entry = build_entry(
            row,
            columns,
            division,
            sex
        )

        entries.append(entry)

    log(f"{len(entries)} lifters processed.")

    if skipped:

        log(f"{skipped} rows skipped.")

    return entries


# ==========================================================
# VALIDATION
# ==========================================================

def validate_entries(entries):
    """
    Perform simple validation before writing.
    """

    if len(entries) == 0:

        error(
            "No lifters were found in the workbook."
        )

    for number, entry in enumerate(entries, start=1):

        #
        # Name is mandatory
        #

        if entry["Name"] == "":

            error(
                f"Entry {number} has no name."
            )

        #
        # Weight class should exist
        #

        if entry["WeightClassLbs"] == "":

            log(
                f"Warning: missing weight class "
                f"for {entry['Name']}"
            )

        #
        # Team is optional
        #

        #
        # Division must exist
        #

        if entry["Division"] == "":

            error(
                "Unable to determine division "
                "from filename."
            )

        #
        # Sex must exist
        #

        if entry["Sex"] == "":

            error(
                "Unable to determine sex "
                "from filename."
            )

    log("Validation complete.")
# ==========================================================
# CSV WRITER
# ==========================================================

def write_entries_csv(entries, output_folder):
    """
    Write entries.csv using the canonical LHSPA column order.
    """

    output_file = output_folder / "entries.csv"

    with open(
        output_file,
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as csvfile:

        writer = csv.DictWriter(
            csvfile,
            fieldnames=OUTPUT_COLUMNS,
            extrasaction="ignore"
        )

        writer.writeheader()

        for entry in entries:
            writer.writerow(entry)

    log(f"Created {output_file}")


# ==========================================================
# MAIN
# ==========================================================

def process_workbook(filename):
    """
    Execute the complete LHSPA workflow.
    """

    source_file = Path(filename)

    if not source_file.exists():
        error(f"File not found:\n{source_file}")

    output_folder = source_file.parent

    #
    # Read workbook
    #

    workbook, worksheet = read_workbook(source_file)

    #
    # Read worksheet rows
    #

    rows = worksheet_to_rows(worksheet)

    #
    # Write original.csv
    #

    write_original_csv(
        rows,
        output_folder
    )

    #
    # Extract first table
    #

    header_row, table_rows = extract_first_table(rows)

    #
    # Build transformed entries
    #

    entries = build_entries(
        header_row,
        table_rows,
        source_file.name
    )

    #
    # Validate output
    #

    validate_entries(entries)

    #
    # Write entries.csv
    #

    write_entries_csv(
        entries,
        output_folder
    )

    log("Transformation complete.")


# ==========================================================
# ENTRY POINT
# ==========================================================

def main():

    if len(sys.argv) != 2:

        print()
        print("Usage:")
        print("    python lhspa_transform.py <workbook.xlsx>")
        print()

        sys.exit(1)

    filename = sys.argv[1]

    try:

        process_workbook(filename)

        print()
        print("=" * 60)
        print("SUCCESS")
        print("=" * 60)
        print()

    except KeyboardInterrupt:

        print()
        print("Cancelled by user.")
        sys.exit(1)

    except Exception as exc:

        error(str(exc))


if __name__ == "__main__":
    main()
    